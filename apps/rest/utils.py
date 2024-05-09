from openpyxl import Workbook, load_workbook


class ExcelWorker:
    def __init__(self, new_file, old_file):
        self.new_file = new_file
        self.old_file = old_file
        self.redundant_items = set()
        self.fetch_duplicate_data()
        self.output_workbook = Workbook()

    def fetch_duplicate_data(self):
        work_book = load_workbook(filename=self.old_file)
        work_sheet = work_book.active
        self.redundant_items = {row[0].value for row in work_sheet.iter_rows(min_row=2)}

    def process_new_file(self):
        output_worksheet = self.output_workbook.active
        main_work_book = load_workbook(filename=self.new_file)
        main_work_sheet = main_work_book.active
        for row in main_work_sheet.iter_rows(min_row=2):
            row_id = row[0].value
            if row_id not in self.redundant_items:
                output_worksheet.append([cell.value for cell in row])

    def get_output_workbook(self):
        return self.output_workbook
